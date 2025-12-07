"""
Excel-де пациент деректерін сақтау және оқу
"""

import openpyxl
from openpyxl import Workbook
from queue_system.patient import Patient
from utils.logger import setup_logger

logger = setup_logger()

FILE_NAME = "patients.xlsx"

class ExcelStorage:
    """Пациенттерді Excel файлында сақтау"""

    def __init__(self):
        try:
            self.wb = openpyxl.load_workbook(FILE_NAME)
            self.sheet = self.wb.active
        except FileNotFoundError:
            self.wb = Workbook()
            self.sheet = self.wb.active
            self.sheet.append(["ID", "Аты", "Жасы", "Дәрігер", "Telegram ID"])
            self.wb.save(FILE_NAME)
            logger.info("Жаңа Excel файл құрылды")

    def save_patient(self, patient: Patient):
        self.sheet.append([patient.id, patient.name, patient.age, patient.doctor, patient.telegram_id])
        self.wb.save(FILE_NAME)
        logger.info("Пациент %s сақталды", patient.name)

    def load_patients(self):
        patients = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            patients.append(Patient(*row))
        return patients

    def remove_patient_by_id(self, patient_id):
        for row in self.sheet.iter_rows(min_row=2):
            if row[0].value == patient_id:
                self.sheet.delete_rows(row[0].row)
                self.wb.save(FILE_NAME)
                logger.info("Пациент %s өшірілді", patient_id)
                return True
        return False
